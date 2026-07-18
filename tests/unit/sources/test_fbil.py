"""Tests for the FBIL source connector."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import httpx
import pytest
import respx

from bonds.config import HttpSettings, Settings
from bonds.http import ThrottledClient
from bonds.models import InstrumentType
from bonds.sources.base import DataUnavailable, SourceError
from bonds.sources.fbil import FbilSource

DATE = dt.date(2026, 7, 10)


@pytest.fixture(autouse=True)
def _no_sleep(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("time.sleep", lambda _s: None)


def _source(tmp_path: Path) -> FbilSource:
    settings = Settings(data_root=tmp_path, http=HttpSettings(min_interval_seconds=0.0))
    return FbilSource(client=ThrottledClient(settings.http), settings=settings)


def test_parse_extracts_records(tmp_path: Path, fbil_gsec_workbook: bytes) -> None:
    records = _source(tmp_path).parse(fbil_gsec_workbook, date=DATE, instrument=InstrumentType.GSEC)
    assert len(records) == 2
    first = records[0]
    assert first.isin == "IN0020160035"
    assert first.coupon == pytest.approx(6.97)
    assert first.maturity_date == dt.date(2026, 9, 6)
    assert first.price == pytest.approx(100.2374)
    assert first.ytm == pytest.approx(5.265)
    assert first.instrument_type is InstrumentType.GSEC
    assert first.quote_date == DATE


def test_parse_ignores_non_isin_rows(tmp_path: Path, fbil_gsec_workbook: bytes) -> None:
    records = _source(tmp_path).parse(fbil_gsec_workbook, date=DATE, instrument=InstrumentType.GSEC)
    assert all(r.isin.startswith("IN") and len(r.isin) == 12 for r in records)


def test_parse_raises_without_header(tmp_path: Path) -> None:
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["no", "isin", "header", "here"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(SourceError, match="ISIN"):
        _source(tmp_path).parse(buf.getvalue(), date=DATE, instrument=InstrumentType.GSEC)


def test_parse_finds_data_sheet_when_not_active(tmp_path: Path) -> None:
    # FBIL sometimes saves the file with a non-data tab (e.g. "Note on FRB & IIB") active while the
    # real G-Sec sheet sits alongside; parse must search all sheets, not trust workbook.active.
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "G-Sec"
    data.append(["ISIN", "Coupon", "Maturity(dd-mmm-yyyy)", "Price(Rs)", "YTM% p.a. (Semi-Annual)"])
    data.append(["IN0020160035", 6.97, dt.datetime(2026, 9, 6), 100.24, 5.27])
    note = wb.create_sheet("Note on FRB & IIB")
    note.append([None, "Note on FRB & IIB", "23-Feb-2023"])
    wb.active = wb.sheetnames.index("Note on FRB & IIB")  # non-data sheet is active
    buf = io.BytesIO()
    wb.save(buf)

    records = _source(tmp_path).parse(buf.getvalue(), date=DATE, instrument=InstrumentType.GSEC)
    assert [r.isin for r in records] == ["IN0020160035"]


def test_parse_non_xlsx_body_is_data_unavailable(tmp_path: Path) -> None:
    # FBIL serves an HTML page (HTTP 200) for dates outside its published range; that must be
    # treated as no-data (SKIPPED), not crash a backfill with BadZipFile.
    html = b"<!DOCTYPE html><html><body>No data</body></html>"
    with pytest.raises(DataUnavailable, match="non-xlsx"):
        _source(tmp_path).parse(html, date=DATE, instrument=InstrumentType.GSEC)


@respx.mock
def test_download_lands_file_and_returns_bytes(tmp_path: Path, fbil_gsec_workbook: bytes) -> None:
    respx.get("https://www.fbil.org.in/wasdm/gsec/downloadPublished").mock(
        return_value=httpx.Response(200, content=fbil_gsec_workbook)
    )
    src = _source(tmp_path)
    content = src.download("gsec", DATE)
    assert content == fbil_gsec_workbook
    landed = tmp_path / "raw" / "fbil" / "gsec" / "2026-07-10.xlsx"
    assert landed.read_bytes() == fbil_gsec_workbook


@respx.mock
def test_download_raises_data_unavailable_on_500(tmp_path: Path) -> None:
    respx.get("https://www.fbil.org.in/wasdm/gsec/downloadPublished").mock(
        return_value=httpx.Response(500)
    )
    with pytest.raises(DataUnavailable):
        _source(tmp_path).download("gsec", DATE)


@respx.mock
def test_download_does_not_retry_holiday_500(tmp_path: Path) -> None:
    # 500 = non-publishing day; it must fail fast, not burn the retry budget.
    route = respx.get("https://www.fbil.org.in/wasdm/gsec/downloadPublished").mock(
        return_value=httpx.Response(500)
    )
    settings = Settings(
        data_root=tmp_path, http=HttpSettings(min_interval_seconds=0.0, max_retries=4)
    )
    source = FbilSource(client=ThrottledClient(settings.http), settings=settings)
    with pytest.raises(DataUnavailable):
        source.download("gsec", DATE)
    assert route.call_count == 1  # not 4


@respx.mock
def test_fetch_valuations_end_to_end(tmp_path: Path, fbil_gsec_workbook: bytes) -> None:
    respx.get("https://www.fbil.org.in/wasdm/gsec/downloadPublished").mock(
        return_value=httpx.Response(200, content=fbil_gsec_workbook)
    )
    records = _source(tmp_path).fetch_valuations("gsec", DATE)
    assert {r.isin for r in records} == {"IN0020160035", "IN0020010081"}


def test_fetch_valuations_rejects_unknown_product(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="unsupported"):
        _source(tmp_path).fetch_valuations("equities", DATE)
